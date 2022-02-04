############## A script that adds a new row to a .numbers file. Used to keep track of the books I've read. ##########

#Importing used libraries
from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side, BORDER_THIN

date = input("On which date did you read the book?") #Get the date of book read from the user
book = input("Which book did you read?") #Get the book read from the user
author = input("Who is the author of the book?") #Get the author of the book from the user
category = input("Which category is the book?") #Get the category of the book from the user
type = input("Paper/E-book/Audiobook?") #Get the type of book read from the user

#Define which data you want to write on the new row
new_row_data = [
    [date, book, author, category, type],
    ]

#Load the Excel Spreadsheet you want to work with.
wb = load_workbook("booksread.xlsx.xlsx")
#Select the first worksheet
ws = wb.worksheets[0]

#Append new rows
for row_data in new_row_data:
    #Append row values
    ws.append(row_data)

#Center all cells
    for col in ws.columns:
        for cell in col:
            alignment_obj = copy(cell.alignment)
            alignment_obj.horizontal = 'center'
            alignment_obj.vertical = 'center'
            cell.alignment = alignment_obj

#Define variable for bold font
bold = Font(bold=True)
#Define variable for thin border
thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000')
)
# Bold text in column A
for cell in ws["A:A"]:
    cell.font = bold
# Set border in column A
for cell in ws['A:A']:
    cell.border = thin_border
# Set border in column B
for cell in ws['B:B']:
    cell.border = thin_border
# Set border in column C
for cell in ws['C:C']:
    cell.border = thin_border
# Set border in column D
for cell in ws['D:D']:
    cell.border = thin_border
# Set border in column E
for cell in ws['E:E']:
    cell.border = thin_border

# Save file
wb.save("booksread.xlsx")

# Print successful message to user
print("Congratulations on reading another book! It was saved successfully!")
